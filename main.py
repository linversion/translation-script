# _*_ coding: utf-8 _*_
import os
import sys
import imp

import openpyxl
import xml.etree.ElementTree as ET
from collections import OrderedDict

def read_excel(path,
               sheet_name='Sheet1',
               start_column=1,
               start_row=1,
               key_column=1,
               end_row=1,
               end_col=1,
               export_direct_to_res=False,
               backup_origin_string_xml=False):
    """read the Excel file

    :param path: the Excel file path
    :type path: string

    :param sheet_name: the sheet name
    :type sheet_name: string

    :param start_column: start column index, start from 1
    :type start_column: int

    :param start_row: start row index, start from 1
    :type start_row: int

    :param key_column: the key name's column index
    :type key_column: int

    :param end_row: end row index
    :type end_row: int

    :param end_col: end column index
    :type end_col: int

    :param export_direct_to_res: whether need to export trans direct to your project's res value folder or not
    :type export_direct_to_res: bool

    :param backup_origin_string_xml: whether need to back up the origin string.xml file
    :type backup_origin_string_xml: bool
    """
    workbook = openpyxl.load_workbook(path)

    sheet = workbook[sheet_name]

    rows = sheet.max_row
    cols = sheet.max_column
    real_end_row = end_row + 1 if end_row != 1 else rows
    real_end_col = end_col + 1 if end_col != 1 else cols + 1
    res_dict = {}
    key_list = []
    read_key = False
    for col in range(start_column, real_end_col):
        language = sheet.cell(1, col).value
        trans_list = []

        for row in range(start_row, real_end_row):
            # key's name, put it in key list
            if not read_key:
                key_name = sheet.cell(row, key_column).value
                key_name = key_name if key_name is not None else 'blank_key_name'
                key_list.append(key_name.replace(' ', '_').lower())

            # get the cell value
            cell_value = sheet.cell(row, col).value
            if cell_value is not None:
                cell_value.replace("'", "\\'").replace("…", "…")
            trans_list.append(cell_value if cell_value is not None else "blank_value")
        read_key = True
        res_dict[language] = trans_list
    # 写入result.xml
    write_result_to_xml(res_dict, key_list)
    # 直接写入项目资源文件
    if export_direct_to_res:
        # need export to res
        for language in res_dict.keys():
            key = language
            print('exporting %s' % language)
            if values_folder_dict.get(key) is not None:
                folder = values_folder_dict[key]
                file_path = '%s%s/strings.xml' % (res_path, folder)
                print('exporting to %s' % folder)
                export_to_xml(file_path, res_dict[language], key_list, backup_origin_string_xml)
    print('finish...')


def export_to_xml(file_path, trans_list, key_list, backup_origin_string_xml):
    # check if need backup
    if backup_origin_string_xml:
        os.rename(file_path, file_path.replace('strings.xml', 'strings-backup.xml'))
    # current strings.xml
    old_dict = parse_android_strings_xml(file_path)
    # replace or add
    for index in range(len(key_list)):
        key = key_list[index]
        value = trans_list[index]
        old_dict[key] = value
    with open(file_path, 'w+', encoding='utf-8') as f:
        f.write("<?xml version=\"1.0\" encoding=\"utf-8\"?>")
        f.write('\n')
        f.write('<resources>')
        f.write('\n')
        for item in old_dict.items():
            key = item[0]
            value = item[1]
            f.write('    ')
            f.write(formatter.format(name=key, str=value))
            f.write('\n')
        f.write('</resources>')
    # with open(file_path, 'a+', encoding='utf-8') as f:
    #
    #     # delete the last </resources> tag
    #     pos = f.tell() - 1
    #     while pos > 0 and f.read(1) != '\n':
    #         pos -= 1
    #         f.seek(pos, os.SEEK_SET)
    #
    #     if pos > 0:
    #         f.seek(pos, os.SEEK_SET)
    #         f.truncate()
    #
    #     f.write('\n')
    #     for index in range(len(key_list)):
    #         key = key_list[index]
    #         value = trans_list[index]
    #         f.write('    ')
    #         f.write(formatter.format(name=key, str=value))
    #         f.write('\n')
    #     f.write('</resources>')


def write_result_to_xml(res_dict, key_list):
    with open('./result.xml', mode='w+', encoding='utf-8') as res_file:
        for (key, value) in res_dict.items():
            language = key
            res_file.write('<!-- {lang} -->\n'.format(lang=language))
            print('write %s' % language)
            for index in range(len(value)):
                lang_value = value[index].strip()
                if lang_value != 'blank_value':
                    if lang_value.startswith('<string'):
                        res_file.write(lang_value)
                    else:
                        res_file.write(formatter.format(name=key_list[index], str=lang_value))

                    res_file.write('\n')
            res_file.write('\n\n')


def excel_column_to_number(column: str) -> int:
    """
    将Excel列名称转换为对应的列号

    参数:
    column (str): Excel列名称（如'A', 'AB', 'XFD'）

    返回:
    int: 对应的列号

    异常:
    ValueError: 如果输入包含非字母字符
    """
    if not column.isalpha():
        raise ValueError("输入必须只包含字母")

    column = column.upper()
    result = 0

    # 从右向左处理每个字符
    for char in column:
        # 验证是否在A-Z范围内
        if not 'A' <= char <= 'Z':
            raise ValueError(f"无效字符 '{char}'")

        # 转换为数字 (A=1, B=2, ..., Z=26)
        value = ord(char) - ord('A') + 1

        # 累加结果，相当于26进制转换
        result = result * 26 + value

    return result


def parse_android_strings_xml(file_path):
    """
    解析 Android 的 strings.xml 文件，返回字典（name -> value）

    参数:
    file_path (str): strings.xml 文件的路径

    返回:
    OrderedDict: 保持原始顺序的字典，键为字符串名称，值为字符串内容
    """
    # 创建有序字典以保持 XML 中的原始顺序
    string_dict = OrderedDict()

    try:
        # 解析 XML 文件
        tree = ET.parse(file_path)
        root = tree.getroot()

        # 遍历所有 <string> 元素
        for string_elem in root.findall('string'):
            # 获取 name 属性值
            name = string_elem.get('name')

            if name:
                # 获取文本内容，处理 CDATA 和转义字符
                text = string_elem.text or ""

                # 处理转义字符（XML 实体）
                # 这里只处理基本转义，如有需要可以扩展
                text = text.replace("&amp;", "&")
                text = text.replace("&lt;", "<")
                text = text.replace("&gt;", ">")
                text = text.replace("&quot;", "\"")
                text = text.replace("&apos;", "'")

                # 添加到字典
                string_dict[name] = text

    except ET.ParseError as e:
        print(f"XML 解析错误: {e}")
    except FileNotFoundError:
        print(f"文件未找到: {file_path}")
    except Exception as e:
        print(f"发生错误: {e}")

    return string_dict


'''
{name} : the key name placeholder
{str} : the translation placeholder
'''
formatter = '<string name=\"{name}\">{str}</string>'
values_folder_dict = {
    '英语': 'values',
    '中文': 'values-zh-rCN',
    # '英文': 'values',
    # '繁体中文': 'values-zh-rTW',
    # '日文': 'values-ja',
    '韩语': 'values-ko',
    # '西班牙语': 'values-es',
    '葡萄牙语（葡萄牙）': 'values-pt',
    '葡萄牙语（巴西）': 'values-pt-rBR',
    # '俄文': 'values-ru',
    '法语': 'values-fr',
    '德语': 'values-de',
    '土耳其语': 'values-tr',
    '意大利语': 'values-it',
    '印尼语': 'values-in',
    '泰语': 'values-th',
    '越南语': 'values-vi',
    '马来语': 'values-ms',
    '菲律宾语': 'values-fil',
    '缅甸语': 'values-my',
    '印度语': 'values-hi'
}
res_path = 'E://Collage/app/src/main/res/'
# res_path = 'res/'
# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    imp.reload(sys)
    read_excel(
        # 'D://collage多语言.xlsx',
        './trans-sample.xlsx',
        sheet_name='Sheet1',
        start_column=excel_column_to_number('B'),
        end_col=excel_column_to_number('C'),
        start_row=2,
        end_row=6,
        key_column=1,
        export_direct_to_res=True,
        backup_origin_string_xml=False
    )

